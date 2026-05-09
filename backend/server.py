from fastapi import FastAPI, APIRouter, HTTPException, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import jwt
import bcrypt
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
SECRET_KEY = os.environ['JWT_SECRET']
ALGORITHM = "HS256"

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

# ============ MODELS ============

class UserCreate(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str

class TokenResponse(BaseModel):
    token: str
    user: dict

class StudentBase(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    age: Optional[int] = None
    enrollment_date: Optional[str] = None
    skill_level: Optional[str] = "Beginner"
    notes: Optional[str] = ""

class StudentCreate(StudentBase):
    pass

class Student(StudentBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class LessonBase(BaseModel):
    student_id: str
    title: str
    description: Optional[str] = ""
    day_of_week: str  # Monday, Tuesday, etc.
    time: str  # HH:MM format
    duration: int = 60  # minutes
    topics_to_teach: Optional[str] = ""

class LessonCreate(LessonBase):
    pass

class Lesson(LessonBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ScheduleOverride(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lesson_id: str
    original_date: str
    new_date: Optional[str] = None
    new_time: Optional[str] = None
    is_cancelled: bool = False
    reason: Optional[str] = ""

# One-time/Manual lesson entry (not recurring)
class OneTimeLessonBase(BaseModel):
    student_id: str
    title: str = "Piano Lesson"
    date: str  # Specific date YYYY-MM-DD
    time: str  # HH:MM format
    duration: int = 60
    topics_to_teach: Optional[str] = ""
    notes: Optional[str] = ""
    is_rescheduled: bool = False  # True if this was rescheduled from a recurring lesson
    original_lesson_id: Optional[str] = None  # Reference to original recurring lesson

class OneTimeLessonCreate(OneTimeLessonBase):
    pass

class OneTimeLesson(OneTimeLessonBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Completed Lesson Model - for tracking actual lesson attendance
class CompletedLessonBase(BaseModel):
    student_id: str
    lesson_id: Optional[str] = None
    date: str  # YYYY-MM-DD format
    duration: int = 60  # minutes
    topics_covered: Optional[str] = ""
    notes: Optional[str] = ""

class CompletedLessonCreate(CompletedLessonBase):
    pass

class CompletedLesson(CompletedLessonBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class FeeBase(BaseModel):
    student_id: str
    amount: float
    due_date: str
    period: str  # e.g., "January 2025"
    status: str = "unpaid"  # paid, unpaid, partial

class FeeCreate(FeeBase):
    pass

class Fee(FeeBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    paid_date: Optional[str] = None
    paid_amount: Optional[float] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class NoteBase(BaseModel):
    student_id: str
    content: str
    lesson_date: Optional[str] = None

class NoteCreate(NoteBase):
    pass

class Note(NoteBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Expense Models
class ExpenseBase(BaseModel):
    name: str
    amount: float
    category: str  # rent, utilities, equipment, supplies, marketing, salary, other
    is_recurring: bool = True
    month: str  # e.g., "January 2025" or "Monthly" for recurring

class ExpenseCreate(ExpenseBase):
    pass

class Expense(ExpenseBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ============ AUTH HELPERS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, email: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7  # 7 days
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============ AUTH ROUTES ============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    hashed = hash_password(user.password)
    
    user_doc = {
        "id": user_id,
        "email": user.email,
        "name": user.name,
        "password": hashed,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, user.email)
    return {"token": token, "user": {"id": user_id, "email": user.email, "name": user.name}}

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(user: UserLogin):
    existing = await db.users.find_one({"email": user.email}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(user.password, existing["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(existing["id"], existing["email"])
    return {"token": token, "user": {"id": existing["id"], "email": existing["email"], "name": existing["name"]}}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

# ============ STUDENT ROUTES ============

@api_router.post("/students", response_model=Student)
async def create_student(student: StudentCreate, current_user: dict = Depends(get_current_user)):
    student_obj = Student(**student.model_dump())
    doc = student_obj.model_dump()
    await db.students.insert_one(doc)
    return student_obj

@api_router.get("/students", response_model=List[Student])
async def get_students(current_user: dict = Depends(get_current_user)):
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    return students

@api_router.get("/students/{student_id}", response_model=Student)
async def get_student(student_id: str, current_user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@api_router.put("/students/{student_id}", response_model=Student)
async def update_student(student_id: str, student: StudentCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Student not found")
    
    update_data = student.model_dump()
    await db.students.update_one({"id": student_id}, {"$set": update_data})
    
    updated = await db.students.find_one({"id": student_id}, {"_id": 0})
    return updated

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.students.delete_one({"id": student_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Also delete related records
    await db.lessons.delete_many({"student_id": student_id})
    await db.fees.delete_many({"student_id": student_id})
    await db.notes.delete_many({"student_id": student_id})
    
    return {"message": "Student deleted successfully"}

# ============ LESSON ROUTES ============

@api_router.post("/lessons", response_model=Lesson)
async def create_lesson(lesson: LessonCreate, current_user: dict = Depends(get_current_user)):
    # Verify student exists
    student = await db.students.find_one({"id": lesson.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    lesson_obj = Lesson(**lesson.model_dump())
    doc = lesson_obj.model_dump()
    await db.lessons.insert_one(doc)
    return lesson_obj

@api_router.get("/lessons", response_model=List[Lesson])
async def get_lessons(student_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    lessons = await db.lessons.find(query, {"_id": 0}).to_list(1000)
    return lessons

@api_router.get("/lessons/{lesson_id}", response_model=Lesson)
async def get_lesson(lesson_id: str, current_user: dict = Depends(get_current_user)):
    lesson = await db.lessons.find_one({"id": lesson_id}, {"_id": 0})
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    return lesson

@api_router.put("/lessons/{lesson_id}", response_model=Lesson)
async def update_lesson(lesson_id: str, lesson: LessonCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.lessons.find_one({"id": lesson_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    update_data = lesson.model_dump()
    await db.lessons.update_one({"id": lesson_id}, {"$set": update_data})
    
    updated = await db.lessons.find_one({"id": lesson_id}, {"_id": 0})
    return updated

@api_router.delete("/lessons/{lesson_id}")
async def delete_lesson(lesson_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.lessons.delete_one({"id": lesson_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lesson not found")
    return {"message": "Lesson deleted successfully"}

# ============ SCHEDULE OVERRIDE ROUTES ============

@api_router.post("/schedule-overrides", response_model=ScheduleOverride)
async def create_schedule_override(override: ScheduleOverride, current_user: dict = Depends(get_current_user)):
    doc = override.model_dump()
    await db.schedule_overrides.insert_one(doc)
    return override

@api_router.get("/schedule-overrides")
async def get_schedule_overrides(lesson_id: Optional[str] = None, date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if lesson_id:
        query["lesson_id"] = lesson_id
    if date:
        query["original_date"] = date
    overrides = await db.schedule_overrides.find(query, {"_id": 0}).to_list(1000)
    return overrides

@api_router.delete("/schedule-overrides/{override_id}")
async def delete_schedule_override(override_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.schedule_overrides.delete_one({"id": override_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Override not found")
    return {"message": "Override deleted successfully"}

# ============ ONE-TIME LESSONS ROUTES ============

@api_router.post("/one-time-lessons", response_model=OneTimeLesson)
async def create_one_time_lesson(lesson: OneTimeLessonCreate, current_user: dict = Depends(get_current_user)):
    # Verify student exists
    student = await db.students.find_one({"id": lesson.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    lesson_obj = OneTimeLesson(**lesson.model_dump())
    doc = lesson_obj.model_dump()
    await db.one_time_lessons.insert_one(doc)
    return lesson_obj

@api_router.get("/one-time-lessons", response_model=List[OneTimeLesson])
async def get_one_time_lessons(student_id: Optional[str] = None, month: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    if month:
        # Filter by month (format: "2025-12" for December 2025)
        query["date"] = {"$regex": f"^{month}"}
    
    lessons = await db.one_time_lessons.find(query, {"_id": 0}).to_list(1000)
    return lessons

@api_router.put("/one-time-lessons/{lesson_id}", response_model=OneTimeLesson)
async def update_one_time_lesson(lesson_id: str, lesson: OneTimeLessonCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.one_time_lessons.find_one({"id": lesson_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    update_data = lesson.model_dump()
    await db.one_time_lessons.update_one({"id": lesson_id}, {"$set": update_data})
    
    updated = await db.one_time_lessons.find_one({"id": lesson_id}, {"_id": 0})
    return updated

@api_router.delete("/one-time-lessons/{lesson_id}")
async def delete_one_time_lesson(lesson_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.one_time_lessons.delete_one({"id": lesson_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lesson not found")
    return {"message": "One-time lesson deleted successfully"}

# Reschedule a recurring lesson occurrence
@api_router.post("/lessons/reschedule")
async def reschedule_lesson(
    lesson_id: str,
    original_date: str,
    new_date: str,
    new_time: Optional[str] = None,
    reason: Optional[str] = "",
    current_user: dict = Depends(get_current_user)
):
    # Get the original lesson
    lesson = await db.lessons.find_one({"id": lesson_id}, {"_id": 0})
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    # Create a schedule override to cancel the original date
    override = ScheduleOverride(
        lesson_id=lesson_id,
        original_date=original_date,
        new_date=new_date,
        new_time=new_time or lesson.get("time"),
        is_cancelled=False,
        reason=reason
    )
    await db.schedule_overrides.insert_one(override.model_dump())
    
    # Create a one-time lesson for the new date
    one_time = OneTimeLesson(
        student_id=lesson["student_id"],
        title=lesson.get("title", "Piano Lesson"),
        date=new_date,
        time=new_time or lesson.get("time"),
        duration=lesson.get("duration", 60),
        topics_to_teach=lesson.get("topics_to_teach", ""),
        is_rescheduled=True,
        original_lesson_id=lesson_id
    )
    await db.one_time_lessons.insert_one(one_time.model_dump())
    
    return {
        "message": "Lesson rescheduled successfully",
        "original_date": original_date,
        "new_date": new_date,
        "new_time": new_time or lesson.get("time")
    }

# ============ COMPLETED LESSONS ROUTES ============

@api_router.post("/completed-lessons", response_model=CompletedLesson)
async def create_completed_lesson(lesson: CompletedLessonCreate, current_user: dict = Depends(get_current_user)):
    # Verify student exists
    student = await db.students.find_one({"id": lesson.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    lesson_obj = CompletedLesson(**lesson.model_dump())
    doc = lesson_obj.model_dump()
    await db.completed_lessons.insert_one(doc)
    return lesson_obj

@api_router.get("/completed-lessons", response_model=List[CompletedLesson])
async def get_completed_lessons(student_id: Optional[str] = None, month: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    if month:
        # Filter by month (format: "2025-12" for December 2025)
        query["date"] = {"$regex": f"^{month}"}
    
    lessons = await db.completed_lessons.find(query, {"_id": 0}).to_list(1000)
    # Sort by date descending
    lessons.sort(key=lambda x: x.get("date", ""), reverse=True)
    return lessons

@api_router.delete("/completed-lessons/{lesson_id}")
async def delete_completed_lesson(lesson_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.completed_lessons.delete_one({"id": lesson_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Completed lesson not found")
    return {"message": "Completed lesson deleted successfully"}

@api_router.get("/completed-lessons/summary/{student_id}")
async def get_student_monthly_summary(student_id: str, month: str, current_user: dict = Depends(get_current_user)):
    """Get a formatted summary of completed lessons for a student in a specific month.
    Month format: YYYY-MM (e.g., 2025-12 for December 2025)
    """
    # Verify student exists
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Get completed lessons for the month
    query = {
        "student_id": student_id,
        "date": {"$regex": f"^{month}"}
    }
    lessons = await db.completed_lessons.find(query, {"_id": 0}).to_list(100)
    lessons.sort(key=lambda x: x.get("date", ""))
    
    # Parse month for display
    try:
        year, month_num = month.split("-")
        month_names = ["January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December"]
        month_name = month_names[int(month_num) - 1]
        period = f"{month_name} {year}"
    except:
        period = month
    
    # Format dates for WhatsApp sharing
    formatted_dates = []
    total_duration = 0
    for lesson in lessons:
        date_str = lesson.get("date", "")
        duration = lesson.get("duration", 60)
        total_duration += duration
        try:
            from datetime import datetime as dt
            date_obj = dt.strptime(date_str, "%Y-%m-%d")
            formatted_dates.append(date_obj.strftime("%d %b %Y (%A)"))
        except:
            formatted_dates.append(date_str)
    
    # Create shareable message
    message = f"🎹 *Grace Music Academy*\n"
    message += f"📋 *Lesson Report - {period}*\n\n"
    message += f"👤 *Student:* {student['name']}\n"
    message += f"📚 *Total Classes:* {len(lessons)}\n"
    message += f"⏱️ *Total Duration:* {total_duration} minutes\n\n"
    message += f"📅 *Completed Lesson Dates:*\n"
    for i, date in enumerate(formatted_dates, 1):
        message += f"{i}. {date}\n"
    
    if not formatted_dates:
        message += "No lessons completed this month.\n"
    
    message += f"\n_Report generated on {datetime.now(timezone.utc).strftime('%d %b %Y')}_"
    
    return {
        "student_name": student["name"],
        "period": period,
        "total_lessons": len(lessons),
        "total_duration": total_duration,
        "dates": formatted_dates,
        "lessons": lessons,
        "shareable_message": message
    }

# ============ FEE ROUTES ============

@api_router.post("/fees", response_model=Fee)
async def create_fee(fee: FeeCreate, current_user: dict = Depends(get_current_user)):
    # Verify student exists
    student = await db.students.find_one({"id": fee.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    fee_obj = Fee(**fee.model_dump())
    doc = fee_obj.model_dump()
    await db.fees.insert_one(doc)
    return fee_obj

@api_router.get("/fees", response_model=List[Fee])
async def get_fees(student_id: Optional[str] = None, status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    if status:
        query["status"] = status
    fees = await db.fees.find(query, {"_id": 0}).to_list(1000)
    return fees

@api_router.get("/fees/{fee_id}", response_model=Fee)
async def get_fee(fee_id: str, current_user: dict = Depends(get_current_user)):
    fee = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    if not fee:
        raise HTTPException(status_code=404, detail="Fee not found")
    return fee

@api_router.put("/fees/{fee_id}", response_model=Fee)
async def update_fee(fee_id: str, fee: FeeCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Fee not found")
    
    update_data = fee.model_dump()
    await db.fees.update_one({"id": fee_id}, {"$set": update_data})
    
    updated = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    return updated

@api_router.put("/fees/{fee_id}/mark-paid")
async def mark_fee_paid(fee_id: str, current_user: dict = Depends(get_current_user)):
    existing = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Fee not found")
    
    await db.fees.update_one(
        {"id": fee_id},
        {"$set": {
            "status": "paid",
            "paid_date": datetime.now(timezone.utc).isoformat(),
            "paid_amount": existing["amount"]
        }}
    )
    
    updated = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    return updated

@api_router.delete("/fees/{fee_id}")
async def delete_fee(fee_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.fees.delete_one({"id": fee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fee not found")
    return {"message": "Fee deleted successfully"}

# ============ NOTE ROUTES ============

@api_router.post("/notes", response_model=Note)
async def create_note(note: NoteCreate, current_user: dict = Depends(get_current_user)):
    # Verify student exists
    student = await db.students.find_one({"id": note.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    note_obj = Note(**note.model_dump())
    doc = note_obj.model_dump()
    await db.notes.insert_one(doc)
    return note_obj

@api_router.get("/notes", response_model=List[Note])
async def get_notes(student_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    notes = await db.notes.find(query, {"_id": 0}).to_list(1000)
    return notes

@api_router.delete("/notes/{note_id}")
async def delete_note(note_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.notes.delete_one({"id": note_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    return {"message": "Note deleted successfully"}

# ============ DASHBOARD STATS ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    total_students = await db.students.count_documents({})
    total_lessons = await db.lessons.count_documents({})
    unpaid_fees = await db.fees.count_documents({"status": "unpaid"})
    paid_fees = await db.fees.count_documents({"status": "paid"})
    
    # Get unpaid fees total
    unpaid_fees_list = await db.fees.find({"status": "unpaid"}, {"_id": 0, "amount": 1}).to_list(1000)
    total_unpaid_amount = sum(fee.get("amount", 0) for fee in unpaid_fees_list)
    
    # Get recent lessons (today's day)
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    today = days[datetime.now(timezone.utc).weekday()]
    today_lessons = await db.lessons.find({"day_of_week": today}, {"_id": 0}).to_list(100)
    
    return {
        "total_students": total_students,
        "total_lessons": total_lessons,
        "unpaid_fees_count": unpaid_fees,
        "paid_fees_count": paid_fees,
        "total_unpaid_amount": total_unpaid_amount,
        "today_lessons": today_lessons,
        "today": today
    }

@api_router.get("/dashboard/upcoming-lessons")
async def get_upcoming_lessons(current_user: dict = Depends(get_current_user)):
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    today_idx = datetime.now(timezone.utc).weekday()
    
    upcoming = []
    for i in range(7):
        day = days[(today_idx + i) % 7]
        lessons = await db.lessons.find({"day_of_week": day}, {"_id": 0}).to_list(100)
        for lesson in lessons:
            student = await db.students.find_one({"id": lesson["student_id"]}, {"_id": 0, "name": 1})
            lesson["student_name"] = student["name"] if student else "Unknown"
            lesson["days_from_now"] = i
            upcoming.append(lesson)
    
    # Sort by days_from_now and time
    upcoming.sort(key=lambda x: (x["days_from_now"], x["time"]))
    return upcoming[:10]

@api_router.get("/dashboard/pending-fees")
async def get_pending_fees(current_user: dict = Depends(get_current_user)):
    fees = await db.fees.find({"status": "unpaid"}, {"_id": 0}).to_list(100)
    for fee in fees:
        student = await db.students.find_one({"id": fee["student_id"]}, {"_id": 0, "name": 1})
        fee["student_name"] = student["name"] if student else "Unknown"
    return fees

# ============ EXPENSE ROUTES ============

@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    expense_obj = Expense(**expense.model_dump())
    doc = expense_obj.model_dump()
    await db.expenses.insert_one(doc)
    return expense_obj

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(category: Optional[str] = None, month: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if category:
        query["category"] = category
    if month:
        query["month"] = month
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(1000)
    return expenses

@api_router.put("/expenses/{expense_id}", response_model=Expense)
async def update_expense(expense_id: str, expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    update_data = expense.model_dump()
    await db.expenses.update_one({"id": expense_id}, {"$set": update_data})
    
    updated = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    return updated

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}

# ============ FINANCE ANALYTICS ============

@api_router.get("/finances/summary")
async def get_finance_summary(current_user: dict = Depends(get_current_user)):
    # Get all fees
    all_fees = await db.fees.find({}, {"_id": 0}).to_list(1000)
    paid_fees = [f for f in all_fees if f.get("status") == "paid"]
    unpaid_fees = [f for f in all_fees if f.get("status") == "unpaid"]
    
    total_income = sum(f.get("amount", 0) for f in paid_fees)
    pending_income = sum(f.get("amount", 0) for f in unpaid_fees)
    
    # Get all expenses
    all_expenses = await db.expenses.find({}, {"_id": 0}).to_list(1000)
    total_expenses = sum(e.get("amount", 0) for e in all_expenses)
    
    # Net profit
    net_profit = total_income - total_expenses
    
    return {
        "total_income": total_income,
        "pending_income": pending_income,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "profit_margin": (net_profit / total_income * 100) if total_income > 0 else 0
    }

@api_router.get("/finances/monthly")
async def get_monthly_finances(current_user: dict = Depends(get_current_user)):
    # Get all paid fees grouped by month
    all_fees = await db.fees.find({"status": "paid"}, {"_id": 0}).to_list(1000)
    all_expenses = await db.expenses.find({}, {"_id": 0}).to_list(1000)
    
    # Create monthly breakdown
    months_data = {}
    
    # Process fees (income)
    for fee in all_fees:
        month = fee.get("period", "Unknown")
        if month not in months_data:
            months_data[month] = {"income": 0, "expenses": 0}
        months_data[month]["income"] += fee.get("amount", 0)
    
    # Process expenses
    for expense in all_expenses:
        month = expense.get("month", "Unknown")
        if expense.get("is_recurring"):
            # Add recurring expense to all months
            for m in months_data:
                months_data[m]["expenses"] += expense.get("amount", 0)
        else:
            if month not in months_data:
                months_data[month] = {"income": 0, "expenses": 0}
            months_data[month]["expenses"] += expense.get("amount", 0)
    
    # Convert to list format for charts
    result = []
    for month, data in months_data.items():
        profit = data["income"] - data["expenses"]
        result.append({
            "month": month,
            "income": data["income"],
            "expenses": data["expenses"],
            "profit": profit
        })
    
    # Sort by month (basic sorting, assumes format "Month Year")
    month_order = ["January", "February", "March", "April", "May", "June", 
                   "July", "August", "September", "October", "November", "December"]
    
    def sort_key(item):
        parts = item["month"].split()
        if len(parts) == 2:
            month_name, year = parts
            try:
                return (int(year), month_order.index(month_name) if month_name in month_order else 12)
            except:
                return (9999, 12)
        return (9999, 12)
    
    result.sort(key=sort_key)
    return result

@api_router.get("/finances/expense-breakdown")
async def get_expense_breakdown(current_user: dict = Depends(get_current_user)):
    all_expenses = await db.expenses.find({}, {"_id": 0}).to_list(1000)
    
    # Group by category
    breakdown = {}
    for expense in all_expenses:
        category = expense.get("category", "other")
        if category not in breakdown:
            breakdown[category] = 0
        breakdown[category] += expense.get("amount", 0)
    
    # Convert to list format for pie chart
    result = [{"category": cat, "amount": amt} for cat, amt in breakdown.items()]
    return result

@api_router.get("/finances/export-excel")
async def export_finances_excel(current_user: dict = Depends(get_current_user)):
    """Export detailed financial report to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Fetch all data
    all_fees = await db.fees.find({}, {"_id": 0}).to_list(1000)
    all_expenses = await db.expenses.find({}, {"_id": 0}).to_list(1000)
    all_students = await db.students.find({}, {"_id": 0}).to_list(1000)
    
    # Create student lookup
    student_map = {s["id"]: s["name"] for s in all_students}
    
    # Create workbook
    wb = Workbook()
    
    # ============ Sheet 1: Monthly Summary ============
    ws1 = wb.active
    ws1.title = "Monthly Summary"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    money_format = '₹#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers1 = ["Month", "Total Income", "Students Paid", "Students Unpaid", "Total Expenses", "Net Profit/Loss", "Status"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Group fees by period
    fees_by_period = {}
    for fee in all_fees:
        period = fee.get("period", "Unknown")
        if period not in fees_by_period:
            fees_by_period[period] = {"paid": [], "unpaid": []}
        if fee.get("status") == "paid":
            fees_by_period[period]["paid"].append(fee)
        else:
            fees_by_period[period]["unpaid"].append(fee)
    
    # Calculate recurring expenses (applied to all months)
    recurring_expenses = sum(e.get("amount", 0) for e in all_expenses if e.get("is_recurring"))
    
    # Non-recurring expenses by month
    non_recurring_by_month = {}
    for e in all_expenses:
        if not e.get("is_recurring"):
            month = e.get("month", "Unknown")
            if month not in non_recurring_by_month:
                non_recurring_by_month[month] = 0
            non_recurring_by_month[month] += e.get("amount", 0)
    
    row = 2
    for period in sorted(fees_by_period.keys()):
        data = fees_by_period[period]
        income = sum(f.get("amount", 0) for f in data["paid"])
        paid_count = len(data["paid"])
        unpaid_count = len(data["unpaid"])
        
        # Get expenses for this period
        period_expenses = recurring_expenses + non_recurring_by_month.get(period, 0)
        profit = income - period_expenses
        status = "Profit" if profit >= 0 else "Loss"
        
        ws1.cell(row=row, column=1, value=period).border = thin_border
        ws1.cell(row=row, column=2, value=income).number_format = money_format
        ws1.cell(row=row, column=2).border = thin_border
        ws1.cell(row=row, column=3, value=paid_count).border = thin_border
        ws1.cell(row=row, column=4, value=unpaid_count).border = thin_border
        ws1.cell(row=row, column=5, value=period_expenses).number_format = money_format
        ws1.cell(row=row, column=5).border = thin_border
        ws1.cell(row=row, column=6, value=profit).number_format = money_format
        ws1.cell(row=row, column=6).border = thin_border
        
        status_cell = ws1.cell(row=row, column=7, value=status)
        status_cell.border = thin_border
        if status == "Profit":
            status_cell.font = Font(color="006400", bold=True)
        else:
            status_cell.font = Font(color="8B0000", bold=True)
        
        row += 1
    
    # Adjust column widths
    for col in range(1, 8):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    
    # ============ Sheet 2: Student Fees Detail ============
    ws2 = wb.create_sheet("Student Fees")
    
    headers2 = ["Student Name", "Period", "Amount", "Due Date", "Status", "Paid Date"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for fee in sorted(all_fees, key=lambda x: (x.get("period", ""), x.get("student_id", ""))):
        ws2.cell(row=row, column=1, value=student_map.get(fee.get("student_id"), "Unknown")).border = thin_border
        ws2.cell(row=row, column=2, value=fee.get("period", "")).border = thin_border
        ws2.cell(row=row, column=3, value=fee.get("amount", 0)).number_format = money_format
        ws2.cell(row=row, column=3).border = thin_border
        ws2.cell(row=row, column=4, value=fee.get("due_date", "")).border = thin_border
        
        status_cell = ws2.cell(row=row, column=5, value=fee.get("status", ""))
        status_cell.border = thin_border
        if fee.get("status") == "paid":
            status_cell.font = Font(color="006400")
        else:
            status_cell.font = Font(color="8B0000")
        
        ws2.cell(row=row, column=6, value=fee.get("paid_date", "")).border = thin_border
        row += 1
    
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    
    # ============ Sheet 3: Fixed Expenses ============
    ws3 = wb.create_sheet("Fixed Expenses")
    
    headers3 = ["Expense Name", "Category", "Amount", "Type", "Period"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for expense in all_expenses:
        ws3.cell(row=row, column=1, value=expense.get("name", "")).border = thin_border
        ws3.cell(row=row, column=2, value=expense.get("category", "")).border = thin_border
        ws3.cell(row=row, column=3, value=expense.get("amount", 0)).number_format = money_format
        ws3.cell(row=row, column=3).border = thin_border
        ws3.cell(row=row, column=4, value="Monthly" if expense.get("is_recurring") else "One-time").border = thin_border
        ws3.cell(row=row, column=5, value=expense.get("month", "")).border = thin_border
        row += 1
    
    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    
    # ============ Sheet 4: Students List ============
    ws4 = wb.create_sheet("Students")
    
    headers4 = ["Name", "Email", "Phone", "Skill Level", "Enrollment Date"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for student in all_students:
        ws4.cell(row=row, column=1, value=student.get("name", "")).border = thin_border
        ws4.cell(row=row, column=2, value=student.get("email", "")).border = thin_border
        ws4.cell(row=row, column=3, value=student.get("phone", "")).border = thin_border
        ws4.cell(row=row, column=4, value=student.get("skill_level", "")).border = thin_border
        ws4.cell(row=row, column=5, value=student.get("enrollment_date", "")).border = thin_border
        row += 1
    
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with date
    filename = f"GraceMusicAcademy_Report_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
